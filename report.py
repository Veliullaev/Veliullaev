import itertools
import pdfkit
from jinja2 import Environment, FileSystemLoader
from openpyxl.styles import Border, Side, Font
from openpyxl.workbook import Workbook
from openpyxl.worksheet import worksheet
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import numpy as np
from xlsx2html import xlsx2html

"""Модуль для преобразования статистических данных в удобочитаемый вид."""

_thin_border = Border(left=Side(style='thin'),
                      right=Side(style='thin'),
                      top=Side(style='thin'),
                      bottom=Side(style='thin'))


def thin_border_all_cells(sheet: worksheet.Worksheet):
    """Метод обводки непустых ячеек.

    Parameters:
        sheet (worksheet.Worksheet): лист Excel
    """
    for i in sheet.__iter__():
        for j in i:
            if j.value == "":
                continue
            j.border = _thin_border


def adjust_columns_width(sheet: worksheet.Worksheet):
    """Метод подгонки шырины ячеек под самую длинную строку в столбце

    Parameters:
        sheet (worksheet.Worksheet): лист Excel
    """
    column_widths = [0] * len(list(sheet.columns))
    for i, column in enumerate(sheet.columns):
        for cell in column:
            if len(str(cell.value)) > column_widths[i]:
                column_widths[i] = len(str(cell.value)) * 1.23
    for i, column_width in enumerate(column_widths, 1):
        sheet.column_dimensions[get_column_letter(i)].width = column_width


def format_worksheets(vac_statistics: worksheet.Worksheet, town_statistics: worksheet.Worksheet):
    """Метод проводит форматирование листов Excel со статистикой по вакансиям и городам

        Parameters:
            vac_statistics (worksheet.Worksheet): лист со статистикой по профессиям
            town_statistics (worksheet.Worksheet): лист со статистикой по городам
    """
    # задание процентовки для статистики по городам (доли вакансий)
    percentages = town_statistics.iter_cols(min_col=5, max_col=5, min_row=2, max_row=town_statistics.max_row)
    for i in percentages:
        for j in i:
            j.style = 'Percent'
            j.number_format = '0.00%'

    # форматирование ячеек для удобной читаемости
    thin_border_all_cells(vac_statistics)
    adjust_columns_width(vac_statistics)
    thin_border_all_cells(town_statistics)
    adjust_columns_width(town_statistics)
    town_statistics.column_dimensions["C"].width = 2

    # задаем полужирный шрифт заголовкам
    vac_statistics["A1"].font = Font(bold=True)
    vac_statistics["B1"].font = Font(bold=True)
    vac_statistics["C1"].font = Font(bold=True)
    vac_statistics["D1"].font = Font(bold=True)
    vac_statistics["E1"].font = Font(bold=True)
    town_statistics["A1"].font = Font(bold=True)
    town_statistics["B1"].font = Font(bold=True)
    town_statistics["D1"].font = Font(bold=True)
    town_statistics["E1"].font = Font(bold=True)


def get_worksheet(ws) -> worksheet.Worksheet:
    """Метод, чтобы PyCharm мог адекватно распознавать листы Excel как листы Excel

    Parameters:
        ws: Лист Excel

    Returns:
        Явный лист Excel.
    """
    return ws


def convert_sheet_to_html_table(sheetname: str):
    """Метод, преобразующий лист Excel в HTML разметку

        Parameters:
            sheetname (str): Excel лист

        Returns:
            Строка HTML-разметки с готовой таблицей.
    """
    out_stream = xlsx2html('report.xlsx', sheet=sheetname)
    out_stream.seek(0)
    years = out_stream.read()
    firstIndex = years.index('<table')
    lastIndex = years.index('</table>')
    return years[firstIndex:lastIndex] + '</table>'


class Report:
    """Класс для создания наглядной статистики в виде Excel, картинки и PDF

        Attributes:
            _profession (str): Имя профессии.
            _salary_dynamic_by_year (dict): Словарь динамики зарплат по годам.
            _vacancy_dynamic_by_year (dict): Словарь динамики числа вакансий по годам.
            _salary_dynamic_profession (dict): Словарь динамики зарплат по годам для конкретной профессии.
            _vacancy_dynamic_profession (dict): Словарь динамики числа вакансий по годам для конкретной профессии..
            _salary_town (dict): Словарь уровня средних зарплат по городам.
            _vacancy_town (dict): Словарь уровня долей вакансий по городам.
    """

    def __init__(self, profession: str, salary_dynamic_by_year: dict, vacancy_dynamic_by_year: dict,
                 salary_dynamic_profession: dict,
                 vacancy_dynamic_profession: dict, salary_town: dict, vacancy_town: dict):
        """Инициализирует класс Report

            Parameters:
                profession (str): Имя профессии.
                salary_dynamic_by_year (dict): Словарь динамики зарплат по годам.
                vacancy_dynamic_by_year (dict): Словарь динамики числа вакансий по годам.
                salary_dynamic_profession (dict): Словарь динамики зарплат по годам для конкретной профессии.
                vacancy_dynamic_profession (dict): Словарь динамики числа вакансий по годам для конкретной профессии..
                salary_town (dict): Словарь уровня средних зарплат по городам.
                vacancy_town (dict): Словарь уровня долей вакансий по городам.
        """
        self._profession = profession
        self._salary_dynamic_by_year = salary_dynamic_by_year
        self._vacancy_dynamic_by_year = vacancy_dynamic_by_year
        self._salary_dynamic_profession = salary_dynamic_profession
        self._vacancy_dynamic_profession = vacancy_dynamic_profession
        self._salary_town = salary_town
        self._vacancy_town = vacancy_town

    def get_profession(self):
        return self._profession

    def get_salary_dynamic_by_year(self):
        return self._salary_dynamic_by_year

    def get_vacancy_dynamic_by_year(self):
        return self._vacancy_dynamic_by_year

    def get_salary_dynamic_profession(self):
        return self._salary_dynamic_profession

    def get_vacancy_dynamic_profession(self):
        return self._vacancy_dynamic_profession

    def get_salary_town(self):
        return self._salary_town

    def get_vacancy_town(self):
        return self._vacancy_town

    def generate_excel(self):
        """Генерирует Excel файл из словарей статистики"""

        data_vacs = [["Год", "Средняя зарплата", format(f"Средняя зарплата - {self._profession}"),
                      "Количество вакансий", f"Количество вакансий - {self._profession}"]]

        for key, value in self._salary_dynamic_by_year.items():
            data_vacs.append([key, value, self._salary_dynamic_profession[key], self._vacancy_dynamic_by_year[key],
                              self._vacancy_dynamic_profession[key]])

        workbook = Workbook()
        vac_statistics: worksheet.Worksheet = get_worksheet(workbook.active)
        vac_statistics.title = "Статистика по годам"
        for i in data_vacs:
            vac_statistics.append(i)
        town_statistics: worksheet.Worksheet = workbook.create_sheet("Статистика по городам")
        data_towns = [["Город", "Уровень зарплат", "", "Город", "Доля вакансий"]]

        # задание ячеек для статистики по городам
        salary_town_trimmed = dict(itertools.islice(self._salary_town.items(), 10))
        vacancy_town_trimmed = dict(itertools.islice(self._vacancy_town.items(), 10))
        for iter1, iter2 in zip(salary_town_trimmed.items(), vacancy_town_trimmed.items()):
            precentage = iter2[1]
            data_towns.append([iter1[0], iter1[1], "", iter2[0], precentage])
        for i in data_towns:
            town_statistics.append(i)

        format_worksheets(vac_statistics, town_statistics)

        workbook.save("report.xlsx")

    def print_statistics(self):
        """Печатает статистику по профессиям"""
        print("Динамика уровня зарплат по годам: ", end="")
        print(self._salary_dynamic_by_year)
        print("Динамика количества вакансий по годам: ", end="")
        print(self._vacancy_dynamic_by_year)
        print("Динамика уровня зарплат по годам для выбранной профессии: ", end="")
        print(self._salary_dynamic_profession)
        print("Динамика количества вакансий по годам для выбранной профессии: ", end="")
        print(self._vacancy_dynamic_profession)
        print("Уровень зарплат по городам (в порядке убывания): ", end="")
        print(dict(itertools.islice(self._salary_town.items(), 10)))
        print("Доля вакансий по городам (в порядке убывания): ", end="")
        print(dict(itertools.islice(self._vacancy_town.items(), 10)))

    def generate_image(self):
        """Генерирует картинку с графиками, отражающими данные о профессиях."""
        fig, ((salaries_year, vacancies_year), (town_salaries, town_vacancies)) = plt.subplots(2, 2)
        self._initialize_salary_stats(salaries_year)
        self._initialize_vacancies_stats(vacancies_year)
        self._initialize_town_salary_stats(town_salaries)
        self._initialize_town_vac_stats(town_vacancies)
        fig.tight_layout()
        fig.savefig("graph.png")

    def _initialize_salary_stats(self, salaries_year):
        """Инициализирует график статистики зарплат"""
        years = list(self._salary_dynamic_by_year.keys())
        average_means = list(self._salary_dynamic_by_year.values())
        profession_means = list(self._salary_dynamic_profession.values())
        x = np.arange(len(years))  # the label locations
        width = 0.35  # the width of the bars
        rects1 = salaries_year.bar(x - width / 2, average_means, width, label='средняя з/п')
        rects2 = salaries_year.bar(x + width / 2, profession_means, width, label=f'з/п {self._profession}')
        salaries_year.set_title('Уровень зарплат по годам', fontsize=8)
        salaries_year.set_xticks(x, years, rotation=90, fontsize=8)
        leg = salaries_year.legend()
        leg.set_in_layout(True)
        salaries_year.grid(axis='y')

    def _initialize_vacancies_stats(self, vacancies_year):
        """Инициализирует график статистики по вакансиям"""
        years = list(self._salary_dynamic_by_year.keys())
        totalVacs = list(self._vacancy_dynamic_by_year.values())
        professionVacs = list(self._vacancy_dynamic_profession.values())
        x = np.arange(len(years))
        width = 0.35
        rects1 = vacancies_year.bar(x - width / 2, totalVacs, width,
                                    label='Количество вакансий')
        rects2 = vacancies_year.bar(x + width / 2, professionVacs, width,
                                    label=f'Количество вакансий\n{self._profession}')
        vacancies_year.set_title("Количество вакансий по годам", fontsize=8)
        vacancies_year.set_xticks(x, years, rotation=90, fontsize=8)
        # vacancies_year.legend()
        leg = vacancies_year.legend()
        leg.set_in_layout(True)
        vacancies_year.grid(axis='y')

    def _initialize_town_salary_stats(self, town_salaries):
        """Инициализирует график статистики по городам"""
        towns_dict = dict(itertools.islice(self._salary_town.items(), 10))
        salaries = list(towns_dict.values())
        towns = list(towns_dict.keys())
        for i, town in enumerate(towns):
            towns[i] = str(town).replace("-", "-\n").replace(" ", "\n")
        y_pos = np.arange(len(towns))
        town_salaries.barh(y_pos, salaries, align='center')
        town_salaries.set_yticks(y_pos, labels=towns, fontsize=6)
        town_salaries.invert_yaxis()
        town_salaries.set_title("Уровень зарплат по городам", fontsize=8)
        town_salaries.grid(axis='x')

    def _initialize_town_vac_stats(self, town_vacs):
        """Инициализирует график статистики по долям вакансий в городах"""
        towns_dict = dict(itertools.islice(self._vacancy_town.items(), 10))
        vacancies_count = list(towns_dict.values())
        vacancies_count.append(1 - sum(vacancies_count))
        towns = list(towns_dict.keys())
        towns.append("Другие")
        labels = towns
        sizes = vacancies_count
        town_vacs.set_title("Доля вакансий по городам", fontsize=8)
        town_vacs.pie(sizes, labels=labels, colors=None, textprops={"fontsize": 6})
        town_vacs.axis('equal')

    def generate_pdf(self):
        """Генерирует pdf файл через конвертацию контента (excel листы и картинка со статистикой) в HTML разметку
        и конвертацию в PDF"""
        table1 = convert_sheet_to_html_table("Статистика по годам")
        table2 = convert_sheet_to_html_table("Статистика по городам")
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")
        pdf_template = template.render({'profession': self._profession, 'table1': table1, 'table2': table2})
        # Здесь поставьте свой путь к wkhtmltopdf.
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, r'report.pdf',
                           configuration=config,
                           options={"enable-local-file-access": ""})


def merge_reports(first: Report, second: Report):
    prof = first.get_profession()
    salary_dynamic_by_year = first.get_salary_dynamic_by_year().update(second.get_vacancy_dynamic_by_year())

    #TODO: надо починить это дело
    #TODO: не надо
    return Report(profession
                  =prof,
                  salary_dynamic_by_year
                  =salary_dynamic_by_year,
                  salary_dynamic_profession
                  =first.get_salary_dynamic_profession(),
                  vacancy_dynamic_profession
                  =first.get_vacancy_dynamic_profession(),
                  vacancy_dynamic_by_year
                  =first.get_vacancy_dynamic_by_year(),
                  vacancy_town={},
                  salary_town={})
